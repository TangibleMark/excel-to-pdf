from openpyxl import load_workbook
from PdfFormFiller import PdfFormFiller


class ExcelReader:
    def __init__(self, excel_workbook):
        self.wb = load_workbook(filename=excel_workbook, data_only=True)

    def get_rows(self, min_row=2, max_row=-1):
        ws = self.wb.active

        if max_row == -1:
            row_range = ws[min_row:ws.max_row]
        else:
            row_range = ws[min_row:max_row]

        for index, row in enumerate(row_range):
            if row[1].value == '#N/A' or None:
                continue
            else:
                # write to pdf
                pdf_filler = PdfFormFiller('REQUEST_FOR_SUPPLEANCE_FORM_.pdf',
                                           f"REQUEST_FOR_SUPPLEANCE_FORM_{row[3].value}_{index}.pdf")

                pdf_filler.write_department()
                pdf_filler.write_sub_emp_number(row[1].value)
                pdf_filler.write_name_of_replacement(row[2].value)
                pdf_filler.write_courses_taught(row[8].value)
                pdf_filler.write_absent_emp_number(row[4].value)
                pdf_filler.write_name_of_absent(row[3].value)
                pdf_filler.write_reason_for_absence('')
                pdf_filler.write_date_of_absence(str(row[7].value).split(" ")[0])
                pdf_filler.write_time_from(str(row[9].value).split("-")[0])
                # pdf_filler.write_time_to(str(row[9].value).split("-")[1])
                pdf_filler.write_hours_worked(row[10].value)


